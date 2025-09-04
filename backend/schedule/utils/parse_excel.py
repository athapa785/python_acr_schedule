import openpyxl
from io import BytesIO

def parse_excel_file(file_data):
    # Load workbook from bytes using openpyxl
    wb = openpyxl.load_workbook(filename=BytesIO(file_data), data_only=True)
    
    sheets_data = []
    # Only process the first worksheet
    first_sheet_name = wb.sheetnames[0] if wb.sheetnames else None
    if first_sheet_name:
        sheet = wb[first_sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=False):
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                    continue

                cell_value = cell.value if hasattr(cell, 'value') else ""
                comment_text = cell.comment.text if hasattr(cell, 'comment') and cell.comment else ""
                
                # Check if the cell has strike-through formatting
                is_strike = False
                if hasattr(cell, 'font') and cell.font:
                    is_strike = getattr(cell.font, 'strike', False)
                
                if comment_text or is_strike:
                    cell_obj = {"value": cell_value if cell_value is not None else ""}
                    if comment_text:
                        cell_obj["comment"] = comment_text
                    if is_strike:
                        cell_obj["strike"] = True
                    row_data.append(cell_obj)
                else:
                    row_data.append(cell_value if cell_value is not None else "")
            data.append(row_data)
        sheets_data.append({
            "name": first_sheet_name,
            "data": data
        })
    return sheets_data